import configparser as cp
import os
import random
import re
import sys
import time
import zipfile
from collections import Counter
from datetime import datetime
import traceback
import urllib.parse
import urllib.request

import logging
import cairosvg
import openpyxl
import pandas as pd
import pytesseract
import requests
from PIL import Image
from fuzzywuzzy import fuzz
from lxml import etree
import colorama


#  使用前须知：
#  1.本版本开发基于python 3.9。 据文档信息，可能使用时py版本要>=3.8
#  2.暂时还没有开发间隔查询功能，网站活着不容易，最好查1000多个停几分钟
#  3.第一次使用前请依次输入以下代码：
#  pip install -i https://pypi.tuna.tsinghua.edu.cn/simple openpyxl
#  pip install -i https://pypi.tuna.tsinghua.edu.cn/simple requests
#  pip install -i https://pypi.tuna.tsinghua.edu.cn/simple pandas
#  pip install -i https://pypi.tuna.tsinghua.edu.cn/simple pytesseract
#  pip install -i https://pypi.tuna.tsinghua.edu.cn/simple cairosvg
#  pip install -i https://pypi.tuna.tsinghua.edu.cn/simple fuzzywuzzy
#  pip install -i https://pypi.tuna.tsinghua.edu.cn/simple python-levenshtein


def enter(data_list):
    zw = input("请输入或拖入你需要导入的数据集")
    sur = os.path.exists(zw)
    data_list = data_list
    if zw == 'y' or zw == 'Y':
        pass
    elif sur:
        data_list.append(zw)
        enter(data_list)
    else:
        print("文件不存在，请重新输入")
        enter(data_list)
    return data_list


def analysis(a, obj5):
    print(a)
    lst = []
    if len(a) == 1:
        if '.tsv' in a[0]:
            with open(a[0], 'r', encoding="utf-8") as f:
                next(f)  # 跳过第一行即可
                for line in f:
                    line = line.strip('\n').split('\t')
                    res = obj5.search(line[0])
                    if res is None:
                        pass
                    else:
                        lst.append("hsa-miR-" + res.group("name"))
                        print(res.group("name"))
        if '.txt' in a[0]:
            f = open(a[0], mode="r", encoding="utf-8")
            gene_lst = f.readlines()
            f.close()
            obj0 = re.compile(r"'(?P<gene_name>.*?)'")
            genes = obj0.finditer(str(gene_lst))
            for gene in genes:
                gene_name = gene.group("gene_name").strip(r"\n")
                lst.append(gene_name)
        source = lst
    else:
        for path in a:
            with open(path, 'r', encoding="utf-8") as f:
                next(f)  # 跳过第一行即可
                for line in f:
                    line = line.strip('\n').split('\t')
                    res = obj5.search(line[6])
                    if res is None:
                        pass
                    else:
                        lst.append("hsa-miR-" + res.group("name"))
        gene_set = dict(Counter(lst))
        source = [key for key, value in gene_set.items() if value > 1]
    print("共" + str(len(source)) + "个交集miRNA")
    print(source)
    return source


# ualcan数据库差异表达查询
def km_expression(a, obj4, output, conf):
    import re
    name = conf[2]["name"]
    folder = conf[2]["folder"]
    cc = str(conf[1]['cc'])
    gap = float(conf[0]['gap'])
    proxy = str(conf[0]['proxy'])
    sheet = output.active
    gene_list = tuple(a[1])
    num = 1
    for gene in gene_list:
        genenam = gene
        ctype = cc
        url = f"http://ualcan.path.uab.edu/cgi-bin/TCGA-miR-Result.pl?genenam={genenam}&ctype={ctype}"
        headers = ["Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) "
                   "Chrome/102.0.0.0 Safari/537.36 ",
                   "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) "
                   "Chrome/103.0.5060.114 Safari/537.36 Edg/103.0.1264.62 ",
                   "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) "
                   "Chrome/103.0.5060.114 Safari/537.36 Edg/103.0.1264.62 ",
                   "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:99.0) Gecko/20100101 Firefox/99.0"]
        header_id = random.randint(0, 3)
        header = {
            "User-Agent": headers[header_id],
        }
        proxies = {
            'http': 'http://' + proxy,
            'https': 'https://' + proxy
        }
        try:
            if proxy:
                resp = requests.get(url=url, headers=header, proxies=proxies, timeout=10)
            else:
                resp = requests.get(url=url, headers=header, timeout=10)
        except BaseException as e:
            if proxy:
                resp = requests.get(url=url, headers=header, proxies=proxies)
                print(e)
            else:
                resp = requests.get(url=url, headers=header)
                print(e)
        resp.encoding = 'utf-8'
        result = obj4.search(resp.text)
        time.sleep(gap)  # 间隔
        print('')
        print("目前差异表达查询至", num, "/", len(gene_list))
        print(gene)
        num = num + 1
        if result is None:
            if '3p' or '5p' in gene:
                print("  未查询到该miRNA成熟体信息")
                miRNA_len = len(gene) - 3
                miRNA = gene[:miRNA_len]
                url1 = f"http://ualcan.path.uab.edu/cgi-bin/TCGA-miR-Result.pl?genenam={miRNA}&ctype={ctype}"
                try:
                    if proxy:
                        html = requests.get(url=url1, headers=header, proxies=proxies, timeout=10)
                    else:
                        html = requests.get(url=url1, headers=header, timeout=10)
                except BaseException as e:
                    if proxy:
                        html = requests.get(url=url1, headers=header, proxies=proxies)
                        print(e)
                    else:
                        html = requests.get(url=url1, headers=header)
                        print(e)
                html.encoding = 'utf-8'
                obj = re.compile(r'No information available')
                sur = obj.search(html.text)
                if sur is None:
                    res = obj4.finditer(html.text)
                    for req in res:
                        value_plus = req.group('value')
                        if value_plus is None or value_plus == 'N/A':
                            index = a[0].get(gene)
                            value_plus1 = 'N/A'
                            sheet.cell(index, 4, value_plus1)
                            output.save(folder + '/' + name + '.xlsx')
                            print("  未查询到该miRNA前体信息")
                        else:
                            index = a[0].get(gene)
                            value_plus1 = value_plus + '（前体）'
                            sheet.cell(index, 4, value_plus1)
                            output.save(folder + '/' + name + '.xlsx')
                            print("  查询到该miRNA前体信息")
                            print("  ", miRNA)
                            print("  ", value_plus)
                else:
                    print("  未查询到该miRNA前体信息")
                    index = a[0].get(gene)
                    value_plus1 = 'N/A'
                    sheet.cell(index, 4, value_plus1)
                    output.save(folder + '/' + name + '.xlsx')
        else:
            res = obj4.finditer(resp.text)
            for re in res:
                value = re.group('value')
                print(value)
                index = a[0].get(gene)
                sheet.cell(index, 4, value)
                output.save(folder + '/' + name + '.xlsx')
    print("差异表达查询完毕")


# ualcan数据库生存曲线查询函数
def km_sur_query(a, output, conf):
    import re
    name = conf[2]["name"]
    folder = conf[2]["folder"]
    proxy = str(conf[0]['proxy'])
    cc = str(conf[1]['cc'])
    gap = float(conf[0]['gap'])
    sheet = output.active
    gene_list = tuple(a[1])
    num = 1
    for gene in gene_list:
        gene_index = tuple([gene])
        gene = gene.lower()
        url2 = f'http://ualcan.path.uab.edu/images/survival-TCGA/miRNA/{cc}-miR-KMinput/Exp/{gene}-KM-Exp.svg'
        headers = ["Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) "
                   "Chrome/102.0.0.0 Safari/537.36 ",
                   "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) "
                   "Chrome/103.0.5060.114 Safari/537.36 Edg/103.0.1264.62 ",
                   "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) "
                   "Chrome/103.0.5060.114 Safari/537.36 Edg/103.0.1264.62 ",
                   "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:99.0) Gecko/20100101 Firefox/99.0"]
        header_id = random.randint(0, 3)
        proxies = {
            'http': 'http://' + proxy,
            'https': 'https://' + proxy
        }
        header = {
            "User-Agent": headers[header_id],
        }
        time.sleep(gap)  # 间隔
        print('')
        print("目前生存分析查询至", num, "/", len(gene_list))
        num = num + 1
        print(gene)
        try:
            if proxy:
                res2 = requests.get(url=url2, headers=header, proxies=proxies, timeout=10)
            else:
                res2 = requests.get(url=url2, headers=header, timeout=10)
        except BaseException as e:
            if proxy:
                res2 = requests.get(url=url2, headers=header, proxies=proxies)
                print(e)
            else:
                res2 = requests.get(url=url2, headers=header)
                print(e)
        content = res2.content
        obj = re.compile(f'svg(.*?)</p>')
        obj2 = re.compile(f'p=(?P<value>.*?\n)')
        res = obj.search(str(content))
        if res is None:
            with open('f.svg', 'wb') as f:
                f.write(content)
                f.close()
            cairosvg.svg2png(url='f.svg', write_to='f.png')
            p_value = pytesseract.image_to_string(Image.open('f.png'), lang='chi_sim+eng')
            os.remove('f.svg')
            os.remove('f.png')
            p_value = p_value.replace(' ', '')
            ps = obj2.finditer(p_value)
            for p in ps:
                value = p.group('value')
                print(value)
                index = a[0].get(gene_index[0])
                sheet.cell(index, 5, value)
                output.save(folder + '/' + name + '.xlsx')
        elif '3p' or '5p' in gene:
            print("  未查询到该miRNA成熟体信息")
            miRNA_len = len(gene) - 3
            miRNA = gene[:miRNA_len]
            url2 = f'http://ualcan.path.uab.edu/images/survival-TCGA/miRNA/{cc}-miR-KMinput/Exp/{miRNA}-KM-Exp.svg'
            try:
                if proxy:
                    html = requests.get(url=url2, headers=header, proxies=proxies, timeout=10)
                else:
                    html = requests.get(url=url2, headers=header, timeout=10)
            except BaseException as e:
                if proxy:
                    html = requests.get(url=url2, headers=header, proxies=proxies)
                    print(e)
                else:
                    html = requests.get(url=url2, headers=header)
                    print(e)
            content2 = html.content
            res = obj.search(str(content2))
            if res is None:
                with open('f.svg', 'wb') as f:
                    f.write(content2)
                    f.close()
                cairosvg.svg2png(url='f.svg', write_to='f.png')
                p_value = pytesseract.image_to_string(Image.open('f.png'), lang='chi_sim+eng')
                os.remove('f.svg')
                os.remove('f.png')
                p_value = p_value.replace(' ', '')
                ps = obj2.finditer(p_value)
                for p in ps:
                    print('  查询到该miRNA前体信息')
                    print("  ", miRNA)
                    value = p.group('value')
                    print("  ", value)
                    index = a[0].get(gene_index[0])
                    value_plus = value + '（前体）'
                    sheet.cell(index, 5, value_plus)
                    output.save(folder + '/' + name + '.xlsx')
            else:
                print("  未查询到该miRNA前体信息")
                index = a[0].get(gene_index[0])
                value_plus = 'N/A'
                sheet.cell(index, 5, value_plus)
                output.save(folder + '/' + name + '.xlsx')


# starbase数据库差异表达查询函数
def bd_query(gene_list, obj, obj2, obj3, output, conf):
    name = conf[2]["name"]
    folder = conf[2]["folder"]
    appear = []  # 防崩溃储存器
    gap = float(conf[0]['gap'])
    proxy = str(conf[0]['proxy'])
    url2 = "https://starbase.sysu.edu.cn/ajaxphp/datatables/diffExpTableResult.php"
    headers = ["Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) "
               "Chrome/102.0.0.0 Safari/537.36 ",
               "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) "
               "Chrome/103.0.5060.114 Safari/537.36 Edg/103.0.1264.62 ",
               "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) "
               "Chrome/103.0.5060.114 Safari/537.36 Edg/103.0.1264.62 ",
               "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:99.0) Gecko/20100101 Firefox/99.0"]
    header_id = random.randint(0, 3)
    header = {
        "User-Agent": headers[header_id],
    }
    proxies = {
        'http': 'http://' + proxy,
        'https': 'https://' + proxy
    }
    num = 1  # 计步器
    index = {}
    for hmd in gene_list:
        data = {
            "gene": hmd,
            "source": "miRNA"
        }
        time.sleep(gap)  # 间隔
        print('')
        print("目前差异表达查询至", num, "/", len(gene_list))
        if proxy:
            html = requests.post(url=url2, headers=header, data=data, proxies=proxies)
        else:
            html = requests.post(url=url2, headers=header, data=data)
        html.encoding = "utf-8"
        print(hmd)
        res = obj.finditer(html.text)  # 解析页面源代码
        sheet = output.active
        for re in res:
            source = re.group("source")
            source = source.replace(" ", "")
            nums = obj2.findall(source)
            value = nums[6]  # 获取COAD
            print(value)
            p = obj3.findall(value)
            num = num + 1
            if "e" in p:  # 判断p值
                gene_name = data.get("gene")
                appear.append(gene_name)
                bd_value = value
                lst = [gene_name, bd_value]
                sheet.append(lst)
                output.save(folder + '/' + name + '.xlsx')
                index[gene_name] = num
            elif float(value) > 0.05 or float(value) == 0:
                continue
            elif 0.05 >= float(value):
                gene_name = data.get("gene")
                appear.append(gene_name)
                bd_value = value
                lst = [gene_name, bd_value]
                sheet.append(lst)
                output.save(folder + '/' + name + '.xlsx')
                index[gene_name] = num
    print('')
    print("差异表达查询完成")
    print("下列miRNA符合条件", appear)
    return index, appear


# starbase数据库生存分析查询函数
def sur_query(a, obj, obj2, obj3, output, conf):
    sheet = output.active
    name = conf[2]["name"]
    folder = conf[2]["folder"]
    gap = float(conf[0]['gap'])
    proxy = str(conf[0]['proxy'])
    gene_list = tuple(a[1])
    appear = a[1]
    num = 1  # 计步器
    for gene in gene_list:
        proxies = {
            'http': 'http://' + proxy,
            'https': 'https://' + proxy
        }
        url = f"https://starbase.sysu.edu.cn/ajaxphp/datatables/survExpTableResult.php"
        headers = ["Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) "
                   "Chrome/102.0.0.0 Safari/537.36 ",
                   "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) "
                   "Chrome/103.0.5060.114 Safari/537.36 Edg/103.0.1264.62 ",
                   "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) "
                   "Chrome/103.0.5060.114 Safari/537.36 Edg/103.0.1264.62 ",
                   "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:99.0) Gecko/20100101 Firefox/99.0"]
        header_id = random.randint(0, 3)
        header = {
            "User-Agent": headers[header_id],
        }
        data = {
            "gene": gene,
            "source": "miRNA"
        }
        time.sleep(gap)  # 间隔
        if proxy:
            resp = requests.post(url=url, headers=header, data=data, proxies=proxies)
        else:
            resp = requests.post(url, headers=header, data=data)
        resp.encoding = "utf-8"
        print('')
        print("目前生存曲线查询至", num, "/", len(gene_list))
        print(gene)
        res = obj.finditer(resp.text)  # 解析页面源代码
        for re in res:
            source = re.group("source")
            source = source.replace(" ", "")
            nums = obj2.findall(source)
            value = nums[5]  # 获取COAD
            print(value)
            p = obj3.findall(value)
            num = num + 1
            index = a[0].get(gene)
            if "e" in p:  # 判断p值
                sheet.cell(index, 3, value)
                output.save(folder + '/' + name + '.xlsx')
            elif float(value) > 0.05 or float(value) == 0:
                sheet.cell(index, 1, '')
                sheet.cell(index, 2, '')
                output.save(folder + '/' + name + '.xlsx')
                appear.remove(gene)
            elif 0.05 >= float(value):
                sheet.cell(index, 3, value)
                output.save(folder + '/' + name + '.xlsx')
    print('')
    print("生存曲线查询完毕")
    return appear


# 论文分析函数
def paper_analaysis(resp, conf):
    analysis = {
        'ACC': ['Adrenocortical Carcinoma', 'ACC', '肾上腺皮质癌'],
        'BLCA': ['BLCA', 'Bladder Urothelial Carcinoma', '膀胱上皮癌', '膀胱癌', '膀胱尿路上皮癌'],
        'BRCA': ['BRCA', 'Breast Invasive Carcinoma', '乳腺癌', '乳腺浸润癌'],
        'CESC': ['CESC', 'Cervical Squamous Cell Carcinoma and Endocervical Adenocarcinoma', '宫颈癌', '宫颈鳞状细胞癌和宫颈腺癌'],
        'CHOL': ['CHOL', 'Cholangiocarcinoma', '胆管癌', '胆管细胞癌'],
        'COAD': ['COAD', 'Colon Adenocarcinoma', '结肠腺癌', '大肠腺癌'],
        'DLBC': ['DLBC', 'Lymphoid Neoplasm Diffuse Large B-cell Lymphoma', '淋巴肿瘤弥漫性大B细胞淋巴瘤'],
        'ESCA': ['ESCA', 'Esophageal Carcinoma', '食管癌'],
        'HNSC': ['HNSC', 'Head and Neck Squamous Cell Carcinoma', '头颈部鳞状细胞癌', '头颈部鳞癌'],
        'KICH': ['KICH', 'Kidney Chromophobe', '肾嫌色细胞瘤'],
        'KIRC': ['KIRC', 'Kidney Renal Clear Cell Carcinoma', '肾透明细胞癌'],
        'KIRP': ['KIRP', 'Kidney Renal Papillary Cell Carcinoma', '肾乳头细胞癌'],
        'LAML': ['LAML', 'Acute Myeloid Leukemia', '急性髓系白血病', '白血病'],
        'LGG': ['LGG', 'Brain Lower Grade Glioma', '脑低级别胶质瘤', '神经胶质瘤'],
        'LIHC': ['LIHC', 'Liver Hepatocellular Carcinoma', '肝细胞癌', '肝癌'],
        'LUAD': ['LUAD', 'Lung Adenocarcinoma', '肺腺癌'],
        'LUSC': ['LUSC', 'Lung Squamous Cell Carcinoma', '肺鳞状细胞癌'],
        'MESO': ['MESO', 'Mesothelioma', '间皮瘤'],
        'OV': ['OV', 'Ovarian Serous Cystadenocarcinoma', '卵巢浆液性囊腺癌', '卵巢癌'],
        'PAAD': ['PAAD', 'Pancreatic Adenocarcinoma', '胰腺癌'],
        'PCPG': ['PCPG', 'Pheochromocytoma and Paraganglioma', '嗜铬细胞瘤和副神经节瘤', '嗜铬细胞瘤', '副神经节瘤'],
        'PRAD': ['PRAD', 'Prostate Adenocarcinoma', '前列腺癌', '前列腺肿瘤'],
        'READ': ['READ', 'Rectum Adenocarcinoma', '直肠癌'],
        'SARC': ['SARC', 'Sarcoma', '肉瘤', '恶性毒瘤'],
        'SKCM': ['SKCM', 'Skin Cutaneous Melanoma', '皮肤黑色素瘤'],
        'STAD': ['STAD', 'Stomach Adenocarcinoma', '胃癌', '胃腺癌'],
        'TGCT': ['TGCT', 'Testicular Germ Cell Tumors', '睾丸生殖细胞瘤', '睾丸癌'],
        'THCA': ['THCA', 'Thyroid Carcinoma', '甲状腺癌', '甲状腺肿瘤'],
        'THYM': ['THYM', 'Thymoma', '胸腺癌', '胸腺肿瘤'],
        'UCEC': ['UCEC', 'Uterine Corpus Endometrial Carcinoma', '子宫体子宫内膜癌'],
        'UCS': ['UCS', 'Uterine Carcinosarcoma', '子宫癌肉瘤'],
        'UVM': ['UVM', 'Uveal Melanoma', '葡萄膜黑色素瘤']
    }
    cc = conf[1]['cc']
    keywords = analysis[cc]
    degree_list = []
    for keyword in keywords:
        degree = fuzz.partial_ratio(keyword, resp)
        degree_list.append(degree)
    degree_list.sort(reverse=True)
    average = sum(degree_list) / len(degree_list)
    final_degree = {'max': degree_list[0], 'average': average}
    return final_degree


#  知网数据库查询
def cnki(a, output, conf, mix):
    import re
    name = conf[2]["name"]
    folder = conf[2]["folder"]
    proxy = str(conf[0]['proxy'])
    sheet = output.active
    gene_list = tuple(a[1])
    num = 1
    gap = float(conf[0]['gap'])
    for gene in gene_list:
        mix = mix
        url = "https://search.cnki.com.cn/Search/Result"
        proxies = {
            'http': 'http://' + proxy,
            'https': 'https://' + proxy
        }
        headers = ["Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) "
                   "Chrome/102.0.0.0 Safari/537.36 ",
                   "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) "
                   "Chrome/103.0.5060.114 Safari/537.36 Edg/103.0.1264.62 ",
                   "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) "
                   "Chrome/103.0.5060.114 Safari/537.36 Edg/103.0.1264.62 ",
                   "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:99.0) Gecko/20100101 Firefox/99.0"]
        header_id = random.randint(0, 3)
        header = {
            "User-Agent": headers[header_id],
        }
        data = {
            "searchType": 'MulityTermsSearch',
            "Content": gene,
            "Order": "1"
        }
        print('')
        print("目前知网文献检索至", num, "/", len(gene_list))
        print(gene)
        time.sleep(gap)
        if proxy:
            resp = requests.post(url=url, headers=header, data=data, proxies=proxies)
        else:
            resp = requests.post(url, headers=header, data=data)
        resp.encoding = 'utf-8'
        obj = re.compile(r'很抱歉')
        res = obj.search(resp.text)
        num = num + 1
        index = a[0].get(gene)
        if res is None:
            print("中国知网中存在关于该miRNA的可疑结果，正在进行可疑度甄别")
            degree = paper_analaysis(resp.text, conf)
            if degree['max'] >= 80:
                print('识别完毕，可疑度为', '\033[1;31mVery High\033[0m', '，不建议进行靶基因筛查')
                mix[gene] = {'cnki_degree': 1, 'index': a[0].get(gene)}
                sheet.cell(index, 6, '\033[1;31mVery High\033[0m')
                output.save(folder + '/' + name + '.xlsx')
            if 30 <= degree['max'] <= 80 or degree['average'] >= 50:
                print('识别完毕，可疑度为', '\033[1;32mHigh\033[0m', '，建议手动确认后进行靶基因筛查')
                mix[gene] = {'cnki_degree': 2, 'index': a[0].get(gene)}
                sheet.cell(index, 6, 'High')
                output.save(folder + '/' + name + '.xlsx')
            if degree['max'] <= 30:
                print('识别完毕，可疑度为', '\033[1;33mlow\033[0m', '，建议手动确认后进行靶基因筛查')
                mix[gene] = {'cnki_degree': 3, 'index': a[0].get(gene)}
                sheet.cell(index, 6, 'Low')
                output.save(folder + '/' + name + '.xlsx')
        else:
            print("中国知网中没有关于该miRNA的可疑结果:", "\033[1;36mNone\033[0m")
            sheet.cell(index, 6, 'None')
            mix[gene] = {'cnki_degree': 4, 'index': a[0].get(gene)}
            output.save(folder + '/' + name + '.xlsx')
    return mix


#  PubMed数据库检索
def pubmed(a, output, conf, mix):
    import re
    name = conf[2]["name"]
    folder = conf[2]["folder"]
    gap = float(conf[0]['gap'])
    sheet = output.active
    gene_list = tuple(a[1])
    num = 1
    proxy = str(conf[0]['proxy'])
    mix = mix
    for gene in gene_list:
        url = f"https://pubmed.ncbi.nlm.nih.gov/?term={gene}"
        headers = ["Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) "
                   "Chrome/102.0.0.0 Safari/537.36 ",
                   "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) "
                   "Chrome/103.0.5060.114 Safari/537.36 Edg/103.0.1264.62 ",
                   "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) "
                   "Chrome/103.0.5060.114 Safari/537.36 Edg/103.0.1264.62 ",
                   "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:99.0) Gecko/20100101 Firefox/99.0"]
        header_id = random.randint(0, 3)
        proxies = {
            'http': 'http://' + proxy,
            'https': 'https://' + proxy
        }
        header = {
            "User-Agent": headers[header_id],
        }
        print('')
        print("目前PubMed文献检索至", num, "/", len(gene_list))
        print(gene)
        time.sleep(gap)
        if proxy:
            resp = requests.get(url=url, headers=header, proxies=proxies)
        else:
            resp = requests.get(url, headers=header)
        resp.encoding = 'utf-8'
        obj = re.compile(r'No results were found')
        res = obj.search(resp.text)
        num = num + 1
        index = a[0].get(gene)
        if res is None:
            print("PubMed中存在关于该miRNA的可疑结果，正在进行可疑度甄别")
            degree = paper_analaysis(resp.text, conf)
            if degree['max'] >= 80:
                print('识别完毕，可疑度为' + '\033[1;31mVery High\033[0m' + '，不建议进行靶基因筛查')
                mix[gene]['pubmed_degree'] = 1
                mix[gene]['index'] = a[0].get(gene)
                sheet.cell(index, 7, 'Very High')
                output.save(folder + '/' + name + '.xlsx')
            if 60 <= degree['max'] <= 80 or degree['average'] >= 30:
                print('识别完毕，可疑度为', '\033[1;32mHigh\033[0m', '，建议手动确认后进行靶基因筛查')
                mix[gene]['pubmed_degree'] = 2
                mix[gene]['index'] = a[0].get(gene)
                sheet.cell(index, 7, 'High')
                output.save(folder + '/' + name + '.xlsx')
            if degree['max'] <= 60:
                print('识别完毕，可疑度为', '\033[1;33mLow\033[0m', '，建议手动确认后进行靶基因筛查')
                mix[gene]['pubmed_degree'] = 3
                mix[gene]['index'] = a[0].get(gene)
                sheet.cell(index, 7, 'low')
                output.save(folder + '/' + name + '.xlsx')
        else:
            print("PubMed中没有关于该miRNA的可疑结果:", "\033[1;36mNone\033[0m")
            sheet.cell(index, 7, 'None')
            mix[gene]['pubmed_degree'] = 4
            mix[gene]['index'] = a[0].get(gene)
            output.save(folder + '/' + name + '.xlsx')
    return mix


# 配置文件解析
def ini():
    filename = 'config.ini'
    wyc = os.path.exists('config.ini')
    if wyc:
        inifile = cp.ConfigParser()
        inifile.read(filename, 'utf-8')
        import re
        obj = re.compile(r"'(?P<data>.*?)'")
        # 解析base
        base = {}
        gap = inifile.get("base", "gap")
        if float(gap) < 0.5:
            print('')
            print('保护数据库从我做起，不要过于频繁请求数据库')
            print('已修改配置文件中间隔为最小时间：' + '\033[1;36m 0.5s\033[0m')
            inifile['base']['gap'] = '0.5'
            with open(filename, 'w') as configfile:
                inifile.write(configfile)
                configfile.close()
        log = inifile.get("base", "log")
        zip = inifile.get('base', 'zip')
        test = inifile.get('base', 'test')
        proxy = inifile.get("base", "proxy")
        if proxy:
            try:
                url = 'http://www.baidu.com/s?wd=beast'
                headers = ["Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) "
                           "Chrome/102.0.0.0 Safari/537.36 ",
                           "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) "
                           "Chrome/103.0.5060.114 Safari/537.36 Edg/103.0.1264.62 ",
                           "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) "
                           "Chrome/103.0.5060.114 Safari/537.36 Edg/103.0.1264.62 ",
                           "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:99.0) Gecko/20100101 Firefox/99.0"]
                header_id = random.randint(0, 3)
                header = {
                    "User-Agent": headers[header_id],
                }
                proxies = {
                    'http': 'http://' + proxy,
                    'https': 'https://' + proxy
                }
                requests.get(url, headers=header, proxies=proxies)
            except BaseException as e:
                print('')
                print('代理服务器无法连接，请重新修改代理服务器IP')
                print(e)
                print(print('\033[1;33m 3\033[0m' + '秒后自动关闭程序'))
                time.sleep(3)
                sys.exit()
            else:
                pass
        e_name = inifile.get("base", "e_name")
        exp_list = ["?", "/", ":", '"', "|", "*", "<", ">", r"'\'"]
        for exp in exp_list:
            if exp in e_name:
                print('')
                print("电子签名错误，请修改后重新进入程序")
                print('\033[1;33m 3\033[0m' + '秒后自动关闭程序')
                time.sleep(3)
                sys.exit()
        input_gse_list = inifile.get('base', 'gse')
        gse = []
        gse_lists = obj.finditer(input_gse_list)
        for gse_list in gse_lists:
            lst = gse_list.group('data')
            if lst == ',' or lst == '':
                pass
            else:
                gse.append(lst)
        base["gap"] = gap
        base["log"] = log
        base["proxy"] = proxy
        base["gse"] = gse
        base["zip"] = zip
        base["e_name"] = e_name
        # 解析query
        query = {}
        cc = inifile.get("query", "cc")
        model = inifile.get("query", "model")
        db = []
        paper = []
        protein = []
        input_db_list = inifile.get('query', 'db')
        save = inifile.get("query", "save")
        venn = inifile.get("query", "venn")
        db_lists = obj.finditer(input_db_list)
        for db_list in db_lists:
            db_lst = db_list.group('data')
            if db_lst == ',':
                pass
            else:
                db.append(db_lst)
        if 'ualcan' in db:
            if test == '1':
                try:
                    requests.get("http://ualcan.path.uab.edu/analysis-mir.html", timeout=8)
                except BaseException as e:
                    print('')
                    print('在测试中发现您的网络不适合请求ualcan数据库（也有可能ualcan噶了或服务器检修）')
                    print('已为您取消查询该数据库')
                    print(e)
                    db.remove('ualcan')
        input_paper_list = inifile.get('query', 'paper')
        paper_lists = obj.finditer(input_paper_list)
        for paper_list in paper_lists:
            paper_lst = paper_list.group('data')
            if paper_lst == ',':
                pass
            else:
                paper.append(paper_lst)
        input_protein_list = inifile.get('query', 'protein')
        protein_lists = obj.finditer(input_protein_list)
        for protein_list in protein_lists:
            protein_lst = protein_list.group('data')
            if protein_lst == ',':
                pass
            else:
                protein.append(protein_lst)
        if 'mirdb' in protein:
            if test == '1':
                try:
                    requests.get("http://mirdb.org/", timeout=10)
                except BaseException as e:
                    print('')
                    print('在测试中发现您的网络不适合请求mirdb数据库（也有可能mirdb噶了或服务器检修）')
                    print('已为您取消查询该数据库')
                    print(e)
                    protein.remove('mirdb')
        if 'mirdip' in protein:
            if test == '1':
                try:
                    requests.get("https://ophid.utoronto.ca/mirDIP/", timeout=10)
                except BaseException as e:
                    print('')
                    print('在测试中发现您的网络不适合请求mirdip数据库（也有可能mirdip噶了或服务器检修）')
                    print('已为您取消查询该数据库')
                    print(e)
                    protein.remove('mirdip')
        if 'mirwalk' in protein:
            if test == '1':
                try:
                    requests.get("http://mirwalk.umm.uni-heidelberg.de/", timeout=10)
                except BaseException as e:
                    print('')
                    print('在测试中发现您的网络不适合请求mirwalk数据库（也有可能mirwalk噶了或服务器检修）')
                    print('已为您取消查询该数据库')
                    print(e)
                    protein.remove('mirwalk')
        if 'targetscan' in protein:
            if test == '1':
                try:
                    requests.get("https://www.targetscan.org/vert_72/", timeout=10)
                except BaseException as e:
                    print('')
                    print('在测试中发现您的网络不适合请求targetscan数据库（也有可能targetscan噶了或服务器检修）')
                    print('已为您取消查询该数据库')
                    print(e)
                    protein.remove('targetscan')
        if 'tarbase' in protein:
            if test == '1':
                try:
                    requests.get("https://dianalab.e-ce.uth.gr/html/diana/web/index.php?r=tarbasev8", timeout=10)
                except BaseException as e:
                    print('')
                    print('在测试中发现您的网络不适合请求tarbase数据库（也有可能tarbase噶了或服务器检修）')
                    print('已为您取消查询该数据库')
                    print(e)
                    protein.remove('tarbase')
        query['db'] = db
        query['paper'] = paper
        query['save'] = save
        query['venn'] = venn
        query['cc'] = cc
        query['protein'] = protein
        query['model'] = model
        if paper and model == "0":
            print('')
            print('文献数据库设置与靶基因数据库检索模式冲突')
            print('\033[1;33m 3\033[0m' + '秒后自动关闭程序')
            time.sleep(3)
            sys.exit()
        elif not paper and model != '0':
            print('')
            print('文献数据库设置与靶基因数据库检索模式冲突')
            print('\033[1;33m 3\033[0m' + '秒后自动关闭程序')
            time.sleep(3)
            sys.exit()
        dt = datetime.now()
        res = dt.strftime('%y-%m-%d %H：%M：%S')
        name_list = {}
        if e_name:
            folder = './output' + '/' + e_name + ' ' + res + ' ' + cc + ' ' + 'gene waver 1.0' + ' output'
            print('')
            print('欢迎' + e_name + '使用本程序')
            print('撒花~')
            time.sleep(0.5)
        else:
            folder = './output' + '/' + res + ' ' + cc + ' ' + 'gene weaver 1.5' + ' output'
        name_list["folder"] = folder
        name = res + ' ' + cc
        name_list['name'] = name
        return base, query, name_list
    else:
        print('')
        print('配置文件损毁或不存在，正在生成中')
        f = open(filename, 'w', encoding='utf-8')
        config = "[base]\n# GSE数据集相对或者绝对地址（格式为在[]中加入,'XXX'即可）(无限制数目)（选填）\ngse = ['','','']\n" \
                 "# 查询间隔时间，建议大于0.5s，不需要添加单位（单位为s）\ngap = 0.5\n" \
                 "\n# 是否生成日志文件，是为1，否为0\nlog = 1\n# 代理服务器ip地址，例如192.168.0.1:80（可不填）\n" \
                 "proxy =\n" \
                 "# 是否将输出文件打包成压缩包（方便课题组在服务器上运行后下载）（选填），是为1，否为0\nzip = 0\n" \
                 r'# 电子签名，方便课题组在服务器或公用电脑上运行时标识身份（选填）(不能出现*:<?/">|\)\ne_name =\n' \
                 '# 测试网站是否可以使用，是为1，否为0（启用时加载配置文件需要10s左右）\ntest = 0\n\n' \
                 "[query]\n# 需要miRNA查询的数据库（格式为在[]中加入,'XXX'即可）\n# 支持数据库：starbase(https://starbase.sysu.edu.cn/)、" \
                 "ualcan（http://ualcan.path.uab.edu/）\n# 因为各方面原因，暂时将starbase数据库作为主数据库，请勿删去’starbase‘\n" \
                 "db = ['starbase','ualcan']\n# 需要查询的癌种（请全部使用大写）（目前仅支持单一癌种查询）" \
                 "\n# 癌种代码详见./cancer.txt文件\ncc = COAD\n# 需要文献检索的数据库（格式为在[]中加入,'XXX'即可）" \
                 "\n# 支持数据库：cnki(https://www.cnki.net/)、PubMed（https://pubmed.ncbi.nlm.nih.gov/）" \
                 "\npaper = ['cnki','pubmed']\n# 需要靶蛋白的数据库（格式为在[]中加入,'XXX'即可）" \
                 "\n# 支持数据库：mirWalk(http://mirwalk.umm.uni-heidelberg.de/)、TargetScan（https://www.targetscan.org" \
                 "/vert_72/）" \
                 "\n# 支持数据库：miRDB(http://mirdb.org/)、mirDIP(https://ophid.utoronto.ca/mirDIP/)" \
                 "\n# 支持数据库：TarBase（https://dianalab.e-ce.uth.gr/html/diana/web/index.php?r=tarbasev8）" \
                 "\nprotein = ['mirdb','mirdip','mirwalk','targetscan','tarbase']" \
                 "\n# 文献重合度检索，在miRNA文献重合度为某等级以上进行靶基因检索\n# 可写参数：4（None）、3（low）、2（high）、1（very high）、0（all）" \
                 "\n# 当不填写文献检索数据库且仍需筛查靶基因时，请使用参数0，其他情况不要使用\n# 文献重合度界定详情请阅读./readme.md文件\nmodel = 3"\
                 "\n# 是否生成靶蛋白韦恩图（查询的数据库数需>=2），是为1，否为0\nvenn = 1\n# 在靶蛋白搜索时是否保存数据库导出的csv文件，是为1，否为0\nsave = 1"
        f.write(config)
        f.close()
        print('')
        print('生成完毕')
        print('请修改配置文件后重新启动程序')
        time.sleep(3)
        sys.exit()


# mirwalk查询函数
def mirwalk_query(conf, gene, target):
    import re
    print('')
    print(gene, " mirwalk开始查询（较慢，请耐心等待）")
    ID = target[gene]['ID']
    proxy = str(conf[0]['proxy'])
    proxies = {
        'http': 'http://' + proxy,
        'https': 'https://' + proxy
    }
    url = f'http://mirwalk.umm.uni-heidelberg.de/human/mirna/{ID}/'
    headers = ["Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) "
               "Chrome/102.0.0.0 Safari/537.36 ",
               "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) "
               "Chrome/103.0.5060.114 Safari/537.36 Edg/103.0.1264.62 ",
               "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) "
               "Chrome/103.0.5060.114 Safari/537.36 Edg/103.0.1264.62 ",
               "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:99.0) Gecko/20100101 Firefox/99.0"]
    header_id = random.randint(0, 3)
    header = {
        "User-Agent": headers[header_id],
    }
    session = requests.session()
    try:
        if proxy:
            session.get(url, headers=header, proxies=proxies, timeout=20)
            html = requests.get(url, headers=header, proxies=proxies, timeout=20)
        else:
            session.get(url, headers=header, timeout=20)
            html = requests.get(url, headers=header, timeout=20)
    except BaseException as e:
        print('cookies失败，正在再次查询')
        time.sleep(6)
        if proxy:
            session.get(url, headers=header, proxies=proxies)
            html = requests.get(url, headers=header, proxies=proxies)
            print(e)
        else:
            session.get(url, headers=header, proxies=proxies)
            html = requests.get(url, headers=header, proxies=proxies)
            print(e)
    obj = re.compile(r'No ID matched your search criteria')
    obj2 = re.compile(r'was not found in our database')
    mirwalk_gene_symbol = []
    if obj.search(html.text) is None and obj2.search(html.text) is None:
        session.get(url, headers=header)
        cookie = session.cookies
        url2 = 'http://mirwalk.umm.uni-heidelberg.de/export/'
        try:
            if proxy:
                resp = requests.get(url2, headers=header, cookies=cookie, proxies=proxies, timeout=20)
            else:
                resp = requests.get(url2, headers=header, cookies=cookie, timeout=20)
        except BaseException as e:
            print('查询失败，正在再次查询')
            time.sleep(6)
            if proxy:
                resp = requests.get(url2, headers=header, cookies=cookie, proxies=proxies, timeout=30)
                print(e)
            else:
                resp = requests.get(url2, headers=header, cookies=cookie, timeout=30)
                print(e)
        resp.encoding = 'utf-8'
        res = resp.content
        name = conf[2]['folder'] + '/' + gene + "/" + "mirWalk" + '.csv'
        with open(name, 'wb') as f:
            f.write(res)
            f.close()
        df = pd.read_csv(name)
        df.set_index('bindingp', inplace=True)
        res = df.loc[1, 'genesymbol']
        for item in res:
            mirwalk_gene_symbol.append(item)
        mirwalk_gene_symbol = list(set(mirwalk_gene_symbol))
        if conf[1]["save"] == '0':
            os.remove(name)
        print(gene, " mirWalk查询完毕")
        print(mirwalk_gene_symbol)
        return mirwalk_gene_symbol
    else:
        print('这个网站没有这个miRNA的相关数据')
        return mirwalk_gene_symbol


# miRDB函数
def mirdb_query(conf, gene):
    import re
    print('')
    print(gene, " miRDB开始查询")
    proxy = str(conf[0]['proxy'])
    url = f'http://mirdb.org/cgi-bin/search.cgi'
    headers = ["Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) "
               "Chrome/102.0.0.0 Safari/537.36 ",
               "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) "
               "Chrome/103.0.5060.114 Safari/537.36 Edg/103.0.1264.62 ",
               "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) "
               "Chrome/103.0.5060.114 Safari/537.36 Edg/103.0.1264.62 ",
               "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:99.0) Gecko/20100101 Firefox/99.0"]
    header_id = random.randint(0, 3)
    header = {
        "User-Agent": headers[header_id],
    }
    proxies = {
        'http': 'http://' + proxy,
        'https': 'https://' + proxy
    }
    data = {
        'species': 'Human',
        'searchBox': gene,
        'submitButton': 'Go',
        'searchType': 'miRNA'
    }
    try:
        if proxy:
            resp = requests.post(url=url, headers=header, proxies=proxies, data=data, timeout=20)
        else:
            resp = requests.post(url, headers=header, data=data, timeout=20)
    except BaseException as e:
        print('查询失败，正在再次查询')
        time.sleep(6)
        if proxy:
            resp = requests.post(url=url, headers=header, proxies=proxies, data=data, timeout=20)
            print(e)
        else:
            resp = requests.post(url, headers=header, data=data, timeout=20)
            print(e)
    resp.encoding = 'utf-8'
    obj = re.compile(r'no Human miRNA is predicted to target symbol')
    mirdb_gene_symbol = []
    if obj.search(resp.text) is None:
        et = etree.HTML(resp.text)
        trs = et.xpath("//tr")
        for tr in trs:
            name = tr.xpath("./td[@width=100]/font/a/text()")
            if name:
                sroce = tr.xpath("./td[@width=65]/p/font/text()")
                if int(sroce[0]) > 60:
                    mirdb_gene_symbol.append(name[0].strip(' '))
        name = conf[2]['folder'] + '/' + gene + "/" + "miRDB" + '.html'
        if conf[1]["save"] == '1':
            with open(name, 'wb') as f:
                f.write(resp.content)
                f.close()
        print(mirdb_gene_symbol)
        return mirdb_gene_symbol
    else:
        print('这个网站没有这个miRNA的相关数据')
        return mirdb_gene_symbol


# mirDIP查询函数
def mirdip_query(conf, gene):
    mirdip_gene_symbol = []
    print('')
    print(gene, " mirDIP开始查询")

    class mirDIP_Http:
        mapScore = {
            'Very High': '0',
            'High': '1',
            'Medium': '2',
            'Low': '3'
        }
        url = "http://ophid.utoronto.ca/mirDIP"
        map = {}  # results will be here

        def __init__(self):
            return

        # unidirectional on genes

        def unidirectionalSearchOnGenes(self, geneSymbols, minimumScore):

            self.sendPost(self.url + "/Http_U", geneSymbols, '', self.mapScore[minimumScore])
            return

        # unidirectional on microrna(s)
        def unidirectionalSearchOnMicroRNAs(self, microRNAs, minimumScore):

            self.sendPost(self.url + "/Http_U", '', microRNAs, self.mapScore[minimumScore])
            return

        # bidirectional
        def bidirectionalSearch(self, geneSymbols, microRNAs, minimumScore, sources, occurrances):
            self.sendPost(self.url + "/Http_B", geneSymbols, microRNAs, self.mapScore[minimumScore], sources,
                          occurrances)
            return

        # .. serve POST request

        def sendPost(self, url_, geneSymbols, microrna, minimumScore, sources='', occurrances='1'):

            params = {
                'genesymbol': geneSymbols,
                'microrna': microrna,
                'scoreClass': minimumScore,
                'dbOccurrences': occurrances,
                'sources': sources}

            params = bytes(urllib.parse.urlencode(params).encode())
            response = ''

            try:
                handler = urllib.request.urlopen(url_, params)
            except Exception:
                traceback.print_exc()
            else:
                self.response = handler.read().decode('utf-8')
                self.makeMap()

            return

        def makeMap(self):

            ENTRY_DEL = 0x01
            KEY_DEL = 0x02

            arr = self.response.split(chr(ENTRY_DEL))

            for str in arr:

                arrKeyValue = str.split(chr(KEY_DEL))
                if len(arrKeyValue) > 1:
                    self.map[arrKeyValue[0]] = arrKeyValue[1]

            return

        def getResulsSize(self):
            if "results_size" in self.map:
                return self.map["results_size"]
            else:
                return ''

        def getResuls(self):
            if "results" in self.map:
                return self.map["results"]
            else:
                return ''

    microRNAs = gene
    # 'Very High', 'High', 'Medium', 'Low'.
    minimumScore = "Very High"
    o = mirDIP_Http()
    o.unidirectionalSearchOnMicroRNAs(microRNAs, minimumScore)
    if o.getResulsSize() == '0':
        print('这个网站没有这个miRNA的相关数据')
        return mirdip_gene_symbol
    else:
        name = conf[2]['folder'] + '/' + gene + "/" + "mirDIP" + '.csv'
        res = o.getResuls().encode(encoding='utf-8')
        a = res.replace(b'\t', b',')
        b = a.replace(b'\r', b'')
        with open(name, 'wb') as f:
            f.write(b)
            f.close()
        df = pd.read_csv(name)
        lst = df.iloc[:, 0]
        for gene_symbol in lst:
            mirdip_gene_symbol.append(gene_symbol)
        print(mirdip_gene_symbol)
        if conf[1]["save"] == 0:
            os.remove(name)
        return mirdip_gene_symbol


# TargetScan查询函数
def targetscan_query(conf, gene):
    import re
    print('')
    print(gene, " TargetScan开始查询")
    magic_book = gene[4:]
    proxy = str(conf[0]['proxy'])
    url = f'https://www.targetscan.org/vert_72/temp/TargetScan7.2__{magic_book}.predicted_targets.xlsx'
    headers = ["Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) "
               "Chrome/102.0.0.0 Safari/537.36 ",
               "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) "
               "Chrome/103.0.5060.114 Safari/537.36 Edg/103.0.1264.62 ",
               "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) "
               "Chrome/103.0.5060.114 Safari/537.36 Edg/103.0.1264.62 ",
               "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:99.0) Gecko/20100101 Firefox/99.0"]
    header_id = random.randint(0, 3)
    header = {
        "User-Agent": headers[header_id],
    }
    proxies = {
        'http': 'http://' + proxy,
        'https': 'https://' + proxy
    }
    try:
        if proxy:
            resp = requests.get(url=url, headers=header, proxies=proxies, timeout=20)
        else:
            resp = requests.get(url, headers=header, timeout=20)
    except BaseException as e:
        print('查询失败，正在再次查询')
        time.sleep(6)
        if proxy:
            resp = requests.get(url=url, headers=header, proxies=proxies, timeout=20)
            print(e)
        else:
            resp = requests.get(url, headers=header, timeout=20)
            print(e)
    resp.encoding = 'utf-8'
    obj = re.compile(r'Not Found')
    res = obj.search(resp.text)
    targetscan_gene_symbol = []
    name = conf[2]['folder'] + '/' + gene + "/" + "TargetScan" + '.xlsx'
    if res is None:
        with open(name, 'wb') as f:
            f.write(resp.content)
            f.close()
        df = pd.read_excel(name, sheet_name=0, index_col=None, usecols=None)
        try:
            lst = df.loc[df['Total context++ score'] < -0.5, 'Target gene']
            for target in lst:
                targetscan_gene_symbol.append(target)
        except BaseException as e:
            e = e
            print('这个网站没有这个miRNA的相关数据')
            return targetscan_gene_symbol
        if conf[1]["save"] == '0':
            os.remove(name)
        print(targetscan_gene_symbol)
        return targetscan_gene_symbol
    else:
        print('这个网站没有这个miRNA的相关数据')
        return targetscan_gene_symbol


# TarBase查询函数
def tarbase_query(conf, gene):
    import re
    print('')
    print(gene, " TarBase开始查询（较慢，请耐心等待）")
    os.mkdir(conf[2]['folder'] + '/' + gene + "/" + "TarBase")
    num = 1
    name = conf[2]['folder'] + '/' + gene + "/" + "TarBase" + '/' + 'TarBase_' + str(num) + '.html'
    proxy = str(conf[0]['proxy'])
    url = f'https://dianalab.e-ce.uth.gr/html/diana/web/index.php?r=tarbasev8/index&miRNAs[]={gene}&genes[' \
          f']=&sources[]=1&sources[]=7&sources[]=9&publication_year=&prediction_score=&sort_field=score&sort_type' \
          f'=DESC&query=1&page={num}'
    headers = ["Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) "
               "Chrome/102.0.0.0 Safari/537.36 ",
               "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) "
               "Chrome/103.0.5060.114 Safari/537.36 Edg/103.0.1264.62 ",
               "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) "
               "Chrome/103.0.5060.114 Safari/537.36 Edg/103.0.1264.62 ",
               "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:99.0) Gecko/20100101 Firefox/99.0"]
    header_id = random.randint(0, 3)
    header = {
        "User-Agent": headers[header_id],
    }
    proxies = {
        'http': 'http://' + proxy,
        'https': 'https://' + proxy
    }
    try:
        if proxy:
            resp = requests.get(url=url, headers=header, proxies=proxies, timeout=30)
        else:
            resp = requests.get(url, headers=header, timeout=30)
    except BaseException as e:
        print('查询失败，正在再次查询')
        time.sleep(6)
        if proxy:
            resp = requests.get(url=url, headers=header, proxies=proxies, timeout=30)
            print(e)
        else:
            resp = requests.get(url, headers=header, timeout=30)
            print(e)
    resp.encoding = 'utf-8'
    obj = re.compile(r'No Results found')
    tarbase_gene_symbol = []
    with open(name, 'wb') as f:
        f.write(resp.content)
        f.close()
    if obj.search(resp.text) is None:
        et = etree.HTML(resp.text)
        trs = et.xpath('//tr[@class="first-level"]')
        jzy = 2
        for tr in trs:
            symbol_name = tr.xpath('./td[1]/text()')
            sroce = tr.xpath('./td/a[@target="_blank"]/text()')
            if sroce:
                if float(sroce[0]) >= 0.5:
                    tarbase_gene_symbol.append(symbol_name[0].replace(' ', ''))
                    jzy = 2
            else:
                jzy = 666
        while jzy < 6:
            url2 = f'https://dianalab.e-ce.uth.gr/html/diana/web/index.php?r=tarbasev8/index&miRNAs[]={gene}&genes[' \
                   f']=&sources[]=1&sources[]=7&sources[]=9&publication_year=&prediction_score=&sort_field=score' \
                   f'&sort_type=DESC&query=1&page={jzy}'
            name1 = conf[2]['folder'] + '/' + gene + "/" + "TarBase" + '/' + 'TarBase_' + str(jzy) + '.html'
            try:
                if proxy:
                    resp = requests.get(url=url2, headers=header, proxies=proxies, timeout=20)
                else:
                    resp = requests.get(url2, headers=header, timeout=20)
            except BaseException as e:
                if proxy:
                    resp = requests.get(url=url2, headers=header, proxies=proxies)
                    print(e)
                else:
                    resp = requests.get(url2, headers=header)
                    print(e)
            resp.encoding = 'utf-8'
            et = etree.HTML(resp.text)
            trs = et.xpath('//tr[@class="first-level"]')
            with open(name1, 'wb') as f:
                f.write(resp.content)
                f.close()
            for tr in trs:
                symbol_name = tr.xpath('./td[1]/text()')
                sroce = tr.xpath('./td/a[@target="_blank"]/text()')
                if sroce:
                    if float(sroce[0]) >= 0.8:
                        tarbase_gene_symbol.append(symbol_name[0].replace(' ', ''))
                    else:
                        jzy = 666
                else:
                    jzy = 666
            jzy += 1
            time.sleep(2)
        print(tarbase_gene_symbol)
        if conf[1]["save"] == '0':
            os.remove(conf[2]['folder'] + '/' + gene + "/" + "TarBase")
        return tarbase_gene_symbol
    else:
        print('这个网站没有这个miRNA的相关数据')
        return tarbase_gene_symbol


# Veen图查询函数：
def venn(conf, gene, gene_symbol, db_list):
    import re
    os.mkdir(conf[2]['folder'] + '/' + gene + "/" + 'Venn')
    list1 = ''
    namelist1 = ''
    list2 = ''
    namelist2 = ''
    list3 = ''
    namelist3 = ''
    list4 = ''
    namelist4 = ''
    list5 = ''
    namelist5 = ''
    if 'mirwalk' in db_list:
        list1 = '\n'.join(gene_symbol['mirwalk'])
        namelist1 = 'mirwalk'
    if 'mirdip' in db_list:
        list2 = '\n'.join(gene_symbol['mirdip'])
        namelist2 = 'mirdip'
    if 'mirdb' in db_list:
        list3 = '\n'.join(gene_symbol['mirdb'])
        namelist3 = 'mirdb'
    if 'targetscan' in db_list:
        list4 = '\n'.join(gene_symbol['targetscan'])
        namelist4 = 'targetsacn'
    if 'tarbase' in db_list:
        list5 = '\n'.join(gene_symbol['tarbase'])
        namelist5 = 'tarbase'
    proxy = str(conf[0]['proxy'])
    url = 'http://bioinformatics.psb.ugent.be/cgi-bin/liste/Venn/calculate_venn.htpl'
    headers = ["Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) "
               "Chrome/102.0.0.0 Safari/537.36 ",
               "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) "
               "Chrome/103.0.5060.114 Safari/537.36 Edg/103.0.1264.62 ",
               "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) "
               "Chrome/103.0.5060.114 Safari/537.36 Edg/103.0.1264.62 ",
               "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:99.0) Gecko/20100101 Firefox/99.0"]
    header_id = random.randint(0, 3)
    header = {
        "User-Agent": headers[header_id],
    }
    proxies = {
        'http': 'http://' + proxy,
        'https': 'https://' + proxy
    }
    data = {
        'version': 'v3',
        'list1': list1,
        'namelist1': namelist1,
        'list2': list2,
        'namelist2': namelist2,
        'list3': list3,
        'namelist3': namelist3,
        'list4': list4,
        'namelist4': namelist4,
        'list5': list5,
        'namelist5': namelist5,
        'shape': 'Symm',
        'fill': 'yes'
    }
    if proxy:
        resp = requests.post(url=url, headers=header, proxies=proxies, data=data)
    else:
        resp = requests.post(url, headers=header, data=data)
    resp.encoding = 'utf-8'
    obj = re.compile(r"href='(?P<path>.*?)' d")
    res = obj.findall(resp.text)
    name = conf[2]['folder'] + '/' + gene + "/" + 'Venn' + '/' + 'Page-' + gene + '.html'
    if resp.text:
        with open(name, "wb") as f:
            f.write(resp.content)
            f.close()
    try:
        zxy = res[1]
    except BaseException as e:
        if proxy:
            resp = requests.post(url=url, headers=header, proxies=proxies, data=data)
        else:
            resp = requests.post(url, headers=header, data=data)
        resp.encoding = 'utf-8'
        obj = re.compile(r"href='(?P<path>.*?)' d")
        res = obj.findall(resp.text)
        print(e)
    else:
        pass
    if resp.text:
        with open(name, "wb") as f:
            f.write(resp.content)
            f.close()
    path = 'http://bioinformatics.psb.ugent.be' + res[1]
    txt = 'http://bioinformatics.psb.ugent.be' + res[2]
    svg = requests.get(path, headers=header)
    svg.encoding = 'utf-8'
    table = requests.get(txt, headers=header)
    table.encoding = 'utf-8'
    svg_name = conf[2]['folder'] + '/' + gene + "/" + 'Venn' + "/" + 'Venn-' + gene + '.svg'
    table_name = conf[2]['folder'] + '/' + gene + "/" + 'Venn' + "/" + 'Table-' + gene + '.txt'
    with open(svg_name, 'wb') as f:
        f.write(svg.content)
        f.close()
    with open(table_name, 'wb') as f:
        f.write(table.content)
        f.close()


# 路由函数
def route(conf, mix, a):
    gene_list = tuple(a[1])
    model = int(conf[1]['model'])
    target = {}
    lst = []
    df = pd.read_csv('databsae/hsa_lists.csv')
    df.set_index('Name', inplace=True)
    gap = float(conf[0]["gap"])
    for gene in gene_list:
        pubmed = mix[gene]['pubmed_degree']
        cnki = mix[gene]['cnki_degree']
        if pubmed > cnki:
            degree = cnki
        else:
            degree = pubmed
        if degree >= model:
            ID = df.loc[gene, 'ID']
            target[gene] = {'index': mix[gene]['index'], 'ID': ID}
            lst.append(gene)
    print("以下miRNA可以使用", lst)
    for gene in lst:
        print('')
        print('开始' + gene + '靶基因筛查')
        gene_symbol = {}
        db_list = []
        os.mkdir(conf[2]["folder"] + '/' + gene)
        if 'mirwalk' in conf[1]['protein']:
            mirwalk_gene_symbol = mirwalk_query(conf, gene, target)
            if mirwalk_gene_symbol:
                gene_symbol['mirwalk'] = mirwalk_gene_symbol
                db_list.append('mirwalk')
            time.sleep(gap)
        if 'mirdb' in conf[1]['protein']:
            mirdb_gene_symbol = mirdb_query(conf, gene)
            if mirdb_gene_symbol:
                gene_symbol['mirdb'] = mirdb_gene_symbol
                db_list.append('mirdb')
            time.sleep(gap)
        if 'targetscan' in conf[1]['protein']:
            targetscan_gene_symbol = targetscan_query(conf, gene)
            if targetscan_gene_symbol:
                gene_symbol['targetscan'] = targetscan_gene_symbol
                db_list.append('targetscan')
            time.sleep(gap)
        if 'mirdip' in conf[1]['protein']:
            mirdip_gene_symbol = mirdip_query(conf, gene)
            if mirdip_gene_symbol:
                gene_symbol['mirdip'] = mirdip_gene_symbol
                db_list.append('mirdip')
            time.sleep(gap)
        if 'tarbase' in conf[1]['protein']:
            tarbase_gene_symbol = tarbase_query(conf, gene)
            if tarbase_gene_symbol:
                gene_symbol['tarbase'] = tarbase_gene_symbol
                db_list.append('tarbase')
            time.sleep(gap)
        intersection = openpyxl.Workbook()
        intersection.save(conf[2]["folder"] + '/' + gene + '/' + 'intersection' + '.xlsx')
        sheet = intersection.active
        num_db = len(conf[1]['protein'])
        num = 1
        while num < num_db:
            sheet.cell(1, num, db_list[num - 1])
            intersection.save(conf[2]["folder"] + '/' + gene + '/' + 'intersection' + '.xlsx')
            step = 2
            for symbol in gene_symbol[db_list[num - 1]]:
                sheet.cell(step, num, symbol)
                step += 1
            intersection.save(conf[2]["folder"] + '/' + gene + '/' + 'intersection' + '.xlsx')
            num += 1
        print('')
        print('靶蛋白文件输出完成')
        if conf[1]['venn'] == '1' and len(conf[1]['protein']) >= 2:
            print('')
            print('韦恩图生成中')
            venn(conf, gene, gene_symbol, db_list)
            print('韦恩图生成完毕')


#  主函数
def main():
    filename = 'config.ini'
    # 获取配置文件
    sst = os.path.exists('./output/')
    if sst:
        pass
    else:
        os.mkdir('./output')
    print('正在读取配置文件')
    conf = ini()
    wyb = os.path.exists('databsae/hsa_lists.csv')
    if wyb:
        pass
    else:
        print('\033[1;31m 基因序列文库损毁\033[0m')
        print('请在https://github.com/zhuerding/gene_weaver中重新下载一份')
        print('自动修复功能开发中……')
        print('\033[1;33m 10\033[0m' + '秒后自动关闭程序')
        time.sleep(10)
        sys.exit()
    print('')
    print('读取配置文件完毕')
    if not conf[0]['gse']:
        data_list = []
        print('')
        print("未识别到配置文件中的GSE数据集地址，手动导入数据集模式启动：")
        print('\033[3;31m 请确保miRNA名位于第一列，且数据集为tsv格式、utf-8编码\033[0m')
        print("当输入" + '\033[1;36m’y‘\033[0m' + "时导入进程结束")
        a = enter(data_list)
        print('')
        print("数据集导入已完成，准备解析")
    else:
        print('')
        print('配置文件中已存在数据集地址')
        a = conf[0]['gse']
        print(a)
        check = input('请确认是否为需要解析的数据集' + '\033[1;36m(y为确认/n为重新导入)\033[0m')
        if check == 'y' or check == 'Y':
            for path in conf[0]['gse']:
                sur = os.path.exists(path)
                if not sur:
                    print('')
                    print(path, "\033[1;31m数据集不存在，导入错误\033[0m")
                    inifile = cp.ConfigParser()
                    inifile.read(filename, 'utf-8')
                    inifile['base']['gse'] = '['']'
                    with open(filename, 'w') as configfile:
                        inifile.write(configfile)
                        configfile.close()
                    print('')
                    print('进入手动导入数据集模式')
                    print('\033[3;31m 请确保基因名位于第一列，且数据集为tsv格式、utf-8编码\033[0m')
                    print("当输入" + '\033[1;36m’y‘\033[0m' + "时导入进程结束")
                    data_list = []
                    a = enter(data_list)
        else:
            print('进入手动导入数据集模式')
            print('\033[3;31m 请确保基因名位于第一列，且数据集为tsv格式、utf-8编码\033[0m')
            print("当输入" + '\033[1;36m’y‘\033[0m' + "时导入进程结束")
            data_list = []
            a = enter(data_list)
            inifile = cp.ConfigParser()
            inifile.read(filename, 'utf-8')
            inifile['base']['gse'] = '['']'
            with open(filename, 'w') as configfile:
                inifile.write(configfile)
                configfile.close()
    obj5 = re.compile(r"hsa-miR-(?P<name>.*)")
    source = analysis(a, obj5)
    # 初始化
    cc = str(conf[1]["cc"])
    obj = re.compile(r"<tr>.*?>" + cc + "(?P<source>.*?)<tr>")
    obj2 = re.compile(r"<td>(.*?)<")
    obj3 = re.compile(r"e")
    obj4 = re.compile(r'Normal-vs-Primary</td>.*?>(?P<value>.*?)</td>')
    # 提取miRNA名
    gene_list = source
    print('')
    print("文件解析完成")
    # 初始化输出文件
    output = openpyxl.Workbook()
    folder = conf[2]['folder']
    name = conf[2]["name"]
    os.mkdir(folder)
    log(folder)
    path = folder
    if conf[0]['log'] == '1':
        make_print_to_file(path)  # 日志记录
    output.save(folder + '/' + name + '.xlsx')
    sheet = output.active
    sheet['A1'] = 'gene name'
    sheet['B1'] = 'expression pvalue(starbase)'
    sheet['C1'] = 'survival analysis pvalue(starbase)'
    sheet['D1'] = 'expression pvalue(ualcan)'
    sheet['E1'] = 'survival analysis pvalue(ualcan)'
    sheet['F1'] = 'cnki'
    sheet['G1'] = 'PubMed'
    output.save(folder + '/' + name + '.xlsx')
    print('\n')
    print("开始查询https://starbase.sysu.edu.cn/数据库")
    print("开始查询差异表达")
    a = bd_query(gene_list, obj, obj2, obj3, output, conf)
    print('')
    print("开始查询生存曲线")
    b = sur_query(a, obj, obj2, obj3, output, conf)
    if not b:
        print('\n')
        print("\033[3;31m没有符合条件的miRNA\033[0m")
        os.remove(folder + '/' + name + '.xlsx')
        print('\033[1;33m 3\033[0m' + '秒后自动关闭程序')
        time.sleep(3)
        sys.exit()
    else:
        print('')
        print("以下miRNA可以使用", b)
        if 'ualcan' in conf[1]["db"]:
            print('')
            print("开始查询http://ualcan.path.uab.edu/数据库")
            print("开始查询差异表达")
            km_expression(a, obj4, output, conf)
            print('')
            print("开始查询生存曲线")
            km_sur_query(a, output, conf)
        mix = {}
        if 'cnki' in conf[1]["paper"]:
            print('')
            print("开始https://www.cnki.net/文献检索")
            mix = cnki(a, output, conf, mix)
        if 'pubmed' in conf[1]['paper']:
            print('')
            print("开始https://pubmed.ncbi.nlm.nih.gov/文献检索")
            pubmed(a, output, conf, mix)
        print('\n')
        print('')
        if mix != {}:
            print('开始靶基因筛查')
            route(conf, mix, a)
        print("正在生成文件")
        if 'ualcan' in conf[1]['db']:
            pass
        else:
            sheet.delete_cols(idx=4, amount=2)
        if 'cnki' in conf[1]['paper']:
            pass
        else:
            sheet.delete_cols(idx=5, amount=1)
        if 'pubmed' in conf[1]['paper']:
            pass
        else:
            sheet.delete_cols(idx=6, amount=1)
        output.save(folder + '/' + name + '.xlsx')
        df = pd.read_excel(folder + '/' + name + '.xlsx', sheet_name='Sheet')
        df1 = df.dropna(subset=['gene name'])
        df1.to_excel(folder + '/' + name + '.xlsx', index=False)
        df2 = df.dropna(axis=1)
        df2.to_excel(folder + '/' + name + '.xlsx', index=False)
        print(df2)
        print("查询完毕" + "\n")
        print("正在保存文件，请不要关闭程序")
        if conf[0]['zip'] == "1":
            month_rank_dir = folder
            zip_file_new = month_rank_dir + '.zip'
            if os.path.exists(month_rank_dir):
                zip = zipfile.ZipFile(zip_file_new, 'w', zipfile.ZIP_DEFLATED)
                for dir_path, dir_names, file_names in os.walk(month_rank_dir):
                    fpath = dir_path.replace(month_rank_dir, '')
                    for filename in file_names:
                        zip.write(os.path.join(dir_path, filename), os.path.join(fpath, filename))
                zip.close()
        time.sleep(2)
        print('保存完毕')


def compressFolder(folderPath, compressPathName):
    zip = zipfile.ZipFile(compressPathName, 'w', zipfile.ZIP_DEFLATED)
    dict = {}
    for path, dirNames, fileNames in os.walk(folderPath):
        fpath = path.replace(folderPath, '')
        for name in fileNames:
            fullName = os.path.join(path, name).decode(encoding='gbk')
            name = fpath + '\\' + name
            zip.write(fullName, name)
            zip.close()


def make_print_to_file(path):
    import os
    import sys
    import datetime

    class Logger(object):
        def __init__(self, filename="Default.log", path=path):
            self.terminal = sys.stdout
            self.log = open(os.path.join(path, filename), "a", encoding='utf8', )

        def write(self, message):
            self.terminal.write(message)
            self.log.write(message)

        def flush(self):
            pass

    fileName = datetime.datetime.now().strftime('day' + '%Y_%m_%d')
    sys.stdout = Logger(fileName + '.log', path=path)
    print(fileName.center(60, '*'))


def log(folder):
    # 设置日志格式
    LOG_FORMAT = "%(asctime)s - %(levelname)s - %(message)s"
    # 设置日期格式
    DATE_FORMAT = "%m/%d/%Y %H:%M:%S %p"
    logging.basicConfig(level=logging.DEBUG, filename=folder + '/bug.log', filemode='a', format=LOG_FORMAT,
                        datefmt=DATE_FORMAT)
    logging.info('This is a info')
    logging.debug('This is a debug')
    logging.warning('This is a warning' + ' ' + time.strftime('%y-%m-%d %H:%M:%S') + ' ' + traceback.format_exc())
    logging.error('This is a error', exc_info=True, stack_info=True, extra={'user': 'Tom', 'ip': '47.98.53.222'})
    logging.critical('This is a critical')
    logging.log(logging.CRITICAL, "This is a critical")


def information():
    print("欢迎使用", "\033[1;36mGene Weaver\033[0m")
    print("查看目前功能及注意事项请移步readme.md")
    print('本程序版本' + '\033[3;31mV1.5.0002\033[0m')
    print('BUG反馈、创意分享请联系\033[3;36mzhuerding@zhuerding.top\033[0m')
    print('\n')
    time.sleep(1)


# 更新函数
def update():
    import re
    url = 'https://github.com/zhuerding/gene_weaver'
    version = '1.5.0002'
    resp = requests.get(url)
    resp.encoding = 'utf-8'
    obj = re.compile(r'>当前版本：V (.*?)  更新日志')
    res = obj.findall(resp.text)
    for re in res:
        if re != version:
            print('有版本更新')
            print('更新日志详见：' + '\033[1;36mhttps://github.com/zhuerding/gene_weaver/blob/master/update.log\033[0m')
            print('源代码地址：' + '\033[1;36mhttps://github.com/zhuerding/gene_weaver\033[0m')
            print('安装包地址：' + '\033[1;36https://pan.baidu.com/s/1tc3yXbjyLs0K-LLsSKYs2A?pwd=pgsc\033[0m')


if __name__ == '__main__':
    colorama.init(autoreset=True)
    t1 = time.time()
    information()
    main()
    t2 = time.time()
    time = t2 - t1
    print('共运行' + f'\033[1;36m{time}\033[0m' + '秒，超过了全国99.99%的用户')
    print('本程序版本' + '\033[3;31mV1.5.0002\033[0m')
    update()
    input('谢谢使用，按任意键退出程序')
